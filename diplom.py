# ПО ИТОГУ: ЗАБЫЛ ЧТО В ОДНОМ ДНЕ НЕСКОЛЬКО ПАР, ПЕРЕДЕЛАТЬ

import pandas as pd
import math
import os
import sys
from openpyxl import Workbook


wb = Workbook()


def main():
    # DEBUG
    sys.argv.append( '.\Задание на ВКР.xls' )

    # загружаем файл, переданный программе в ее аргументах
    file = load_file()

    # просим пользователя выбрать лист для работы программы
    selected_list = select_list(file)
    # на этом моменте получили целый лист, готовый к парсингу (selected_list)

    # парсим лист в список пар
    lessons = list_to_lessons(selected_list)

    # формируем расписание
    timetables = lessons_to_timetables(lessons)

    xl_print_timetables(timetables)

    #print( timetables['КСП224']['even'][0] )
    #print( timetables['КСП225']['even'][0] )
    f = open("result.txt", "w")
    f.write( str(timetables) )
    f.close()

def beauty_print_timetables(timetables):
    for timetable in timetables:
        print(timetable)
        timetable = timetables[timetable]
        print("Четная неделя")
        for idx, day in enumerate(timetable['even']):
            print(f'День {idx+1}:')
            for i, lesson in enumerate(day):
                if lesson == '':
                    continue
                print(f'{i}. {lesson["lesson"]}')
        print()
        print("Нечетная неделя")
        for idx, day in enumerate(timetable['odd']):
            print(f'День {idx+1}:')
            for i, lesson in enumerate(day):
                if lesson == '':
                    continue
                print(f'{i}. {lesson["lesson"]}')
        print()
        print()
        print()

def xl_print_timetables(timetables):
    for timetable in timetables:
        ws = wb.create_sheet(timetable)

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50

        timetable = timetables[timetable]
        ws.cell(row=1,column=1).value = 'Чётная неделя'
        for idx, day in enumerate(timetable['even']):
            for i, lesson in enumerate(day):
                if lesson == '':
                    continue
                else:
                    ws.cell(row=i+2,column=idx+1).value = f'{i+1}. {lesson["lesson"]}'
        ws.cell(row=14,column=1).value = 'Чётная неделя'
        for idx, day in enumerate(timetable['odd']):
            for i, lesson in enumerate(day):
                if lesson == '':
                    continue
                else:
                    ws.cell(row=i+15,column=idx+1).value = f'{i+1}. {lesson["lesson"]}'
    wb.save("sample.xlsx")
    wb.close()

def lessons_to_timetables(lessons):
    timetables = {} # массив расписаний, 1 элемент - 1 расписание для группы

    # формируем список групп в timetables
    for lesson in lessons:
        # добавляем группу в timetables если ее там еще нет
        if (not lesson['group'] in timetables):
            #print("Добавляем группу " + lesson['group'])
            timetables[lesson['group']] = {
                'even': [[], [], [], [], [], [], []], # четная неделя
                'odd': [[], [], [], [], [], [], []] # нечетная неделя
            }

    # в дни добавляем пустые предметы
    for group in timetables:
        timetable = timetables[group]
        for week in timetable:
            week = timetable[week]
            for day in week:
                for i in range(10):
                    day.append('')
    
    # создаем массив пар, которые не удалось установить
    failed_lessons = []
    # работаем отдельно по каждой группе
    for group in timetables.keys():
        week = 'even' # начинаем с четной недели
        
        # получаем список запланированных занятий у этой группы
        lessons_of_group = get_lessons_of_group(lessons, group)

        # формируем расписание
        fill_timetable(timetables, group, lessons_of_group)
    
    return timetables

'''
Алгоритм заполнения расписания:
1. Проверить, что пара свободна
2. Внести пару
3. Перейти на следующий день
Повторять пока не кончатся пары
'''
def fill_timetable(timetables, group, week_lessons):
    day = 0 # начинаем с понедельника
    lesson_num = 0 # начинаем с 1 пары
    week = 'even' #  начинаем с четной недели

    while len(week_lessons) > 0:
        lesson = week_lessons[0]

        # проверяем, что на эту пару не установлена другая
        if timetables[group][week][day][lesson_num] == '':
            # проверяем свободен ли урок
            if is_lesson_free(timetables, lesson['lesson'], week == 'even', day, lesson_num):
                # размещение пары на своем месте
                timetables[group][week][day][lesson_num] = lesson

                # удаляем пару из week_lessons
                if (lesson['hours'] < 3):
                    week_lessons.remove(lesson)
                else:
                    lesson['hours'] -= 2

                # возвращаемся на первую пару первого дня
                day = 0
                lesson_num = 0
            else:
                # пробуем поставить в другое место
                day += 1
        
        day += 1
        if (day >= 5):
            if (week == 'even'):
                week = 'odd'
                day = 0
            else:
                week = 'even'
                day = 0
                lesson_num += 1

# из списка общего списка занятий возвращает занатия только для определенной группы
def get_lessons_of_group(lessons, target_group):
    result = []
    for lesson in lessons:
        if lesson['group'] == target_group:
            result.append(lesson)
    return result

# функция проверяет, свободен ли определенный предмет на конкретную неделю на конкретной паре
def is_lesson_free(timetables, lesson_name, is_week_even, day, lesson_num) -> bool:
    # проходимся по всем расписаниям
    for timetable in timetables:
        timetable = timetables[timetable]
        if is_week_even:
            week = 'even'
        else:
            week = 'odd'
        try:
            # если эта пара уже занята
            if timetable[week][day][lesson_num]['lesson'] == lesson_name:
                return False
            else:
                continue
        except:
            continue
    return True

# функция проверяет, свободена конкретная пара, т.е. можно ли туда вотнкуть пару или она занята
#def is_lesson_free2():
#    timetables[group][week][day][lesson_num]

def list_to_lessons(lessons_list):
    # забираем значения таблицы
    table = lessons_list.values

    # ищем номер первой строки с парой
    first_row_num = None
    for idx, row in enumerate(table):
        temp = str(row[4])
        is_nan = temp == 'nan'
        if is_nan:
            continue
        if temp.isnumeric():
            first_row_num = idx
            break

    # от первой строки парсим все последующие, получая в результате массив с данными по парам
    lessons = []
    for i in range(first_row_num, len(table)):
        if str(table[i][0]) == 'nan':
            table[i][0] = ''
        new_lesson = {
            'teacher': str(table[i][0]),
            'lesson': str(table[i][1]),
            'group': str(table[i][2]),
            'hours': int(table[i][4])
        }
        lessons.append(new_lesson)

    return lessons

def select_list(file):
    # ввод листа от пользователя
    # selected_list = 0
    # if (len(xl.sheet_names) > 1):
    #     print("Обнаружено несколько листов. Выберите необходимый лист:")
    #     # Печатаем название листов в данном файле
    #     for idx, list_name in enumerate(xl.sheet_names):
    #         print(f"{idx+1}. {list_name}")
    #     selected_list = input("Введите число: ")
    selected_list = 2
    try:
        selected_list = int(selected_list)
        selected_list -= 1
        # Загрузить лист в DataFrame по его имени: df1
        selected_list = file.parse(file.sheet_names[selected_list])
    except:
        print("Лист некорректен!")
    return selected_list

def load_file():
    # проверяем наличие переданного файла
    if (len(sys.argv) < 2):
        print("Ошибка: запуск без аргументов!")
        exit()
    else:
        in_file = sys.argv[1]

    # Загружаем spreadsheet в объект pandas
    file = pd.ExcelFile(in_file)
    return file

def debug_message(text):
    print(text)

if __name__ == '__main__':
    main()